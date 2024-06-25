import openpyxl as xl
from openpyxl.chart import BarChart,Reference
def process_workbook(filename):
   wb = xl.load_workbook(filename)#enter excel filename
   sheet=wb['Sheet1']
   cell=sheet['a1']
   cell=sheet.cell(1,1)
   for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,4)
        discount_price=cell.value*0.9 #removing 10% of each product price
        corrected_price=sheet.cell(row,5)
        corrected_price.value=discount_price
        Values=Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=5,
              max_col=5)
        chart=BarChart()
        chart.add_data(Values)
        sheet.add_chart(chart,'H2')
        wb.save(filename)


